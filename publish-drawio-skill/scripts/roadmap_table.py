#!/usr/bin/env python3
"""Import the canonical XLSX or long-form CSV into deterministic roadmap.v2 YAML."""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import sys
from collections import defaultdict
from pathlib import Path

import yaml

from roadmap_validate import validate_document
from validation_common import ValidationReport


COMPUTED_COLUMNS = {"previous_planned_date", "shift_days", "cumulative_shift_days", "shift_state"}
SHEETS = ("Settings", "Lanes", "Tasks", "MilestoneHistory", "Dependencies", "Outcomes")


def clean(value):
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).strip()


def formula_like(value) -> bool:
    return isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@"))


def table_rows(path: Path, report: ValidationReport):
    if path.suffix.lower() == ".xlsx":
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for XLSX import; install requirements.txt") from exc
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=True, keep_links=False)
        result = {}
        for sheet_name in SHEETS:
            if sheet_name not in workbook.sheetnames:
                report.add("table", "error", "roadmap.table.sheet.missing", f"/{sheet_name}", f"required sheet {sheet_name!r} is missing")
                result[sheet_name] = []
                continue
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows()
            try:
                header_cells = next(iterator)
            except StopIteration:
                report.add("table", "error", "roadmap.table.header.missing", f"/{sheet_name}", "sheet is empty")
                result[sheet_name] = []
                continue
            headers = [clean(cell.value).lower() for cell in header_cells]
            rows = []
            for row_number, cells in enumerate(iterator, 2):
                row = {}
                user_nonempty = False
                for header, cell in zip(headers, cells):
                    if not header:
                        continue
                    value = cell.value
                    if header not in COMPUTED_COLUMNS and value not in (None, ""):
                        user_nonempty = True
                    if header not in COMPUTED_COLUMNS and (cell.data_type == "f" or formula_like(value)):
                        report.add("table", "error", "roadmap.table.formula.unsafe", f"/{sheet_name}/{row_number}/{header}", "formulas are allowed only in computed columns")
                    row[header] = clean(value)
                # Pre-filled formulas in otherwise blank template rows are not
                # user data and must not produce missing-id findings.
                if user_nonempty:
                    row["_row"] = row_number
                    rows.append(row)
            result[sheet_name] = rows
        workbook.close()
        return result

    if path.suffix.lower() != ".csv":
        raise ValueError("input must have .xlsx or .csv extension")
    result = {name: [] for name in SHEETS}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames or "entity_type" not in {name.strip().lower() for name in reader.fieldnames}:
            report.add("table", "error", "roadmap.table.csv.entity_type.missing", "/header", "CSV requires an entity_type column")
            return result
        for row_number, raw in enumerate(reader, 2):
            row = {clean(key).lower(): clean(value) for key, value in raw.items() if key is not None}
            if not any(value for field, value in row.items() if field not in COMPUTED_COLUMNS):
                continue
            for field, value in row.items():
                if field not in COMPUTED_COLUMNS and formula_like(value):
                    report.add("table", "error", "roadmap.table.formula.unsafe", f"/{row_number}/{field}", "spreadsheet formula-like values are not accepted in CSV input")
            entity = row.get("entity_type", "").lower()
            mapping = {
                "setting": "Settings", "lane": "Lanes", "task": "Tasks",
                "milestone_revision": "MilestoneHistory", "dependency": "Dependencies", "outcome": "Outcomes",
            }
            if entity not in mapping:
                report.add("table", "error", "roadmap.table.csv.entity_type.invalid", f"/{row_number}/entity_type", f"unknown entity_type {entity!r}")
                continue
            if entity == "setting":
                row["key"], row["value"] = row.get("setting_key", ""), row.get("setting_value", "")
            row["_row"] = row_number
            result[mapping[entity]].append(row)
    return result


def split_ids(value):
    return sorted({part.strip() for part in clean(value).replace(";", "|").replace(",", "|").split("|") if part.strip()})


def boolean(value):
    return clean(value).lower() in {"1", "true", "yes", "y", "да"}


def integer(value, report, path, code):
    try:
        return int(value)
    except (TypeError, ValueError):
        report.add("table", "error", code, path, f"expected integer, got {value!r}")
        return value


def iso_date(value, report, path):
    value = clean(value)
    try:
        return dt.date.fromisoformat(value).isoformat()
    except (TypeError, ValueError):
        report.add("table", "error", "roadmap.table.date.invalid", path, f"expected ISO date YYYY-MM-DD, got {value!r}")
        return value


def compact(mapping):
    return {key: value for key, value in mapping.items() if value not in (None, "", [])}


def import_model(path: Path, strict=False):
    report = ValidationReport(schema_version=2)
    tables = table_rows(path, report)
    settings = {}
    for row in tables["Settings"]:
        key = row.get("key", "")
        if not key:
            continue
        if key in settings:
            report.add("table", "error", "roadmap.table.setting.duplicate", f"/Settings/{row['_row']}/key", f"duplicate setting {key!r}")
        settings[key] = row.get("value", "")
    version = integer(settings.get("schema_version", "2"), report, "/Settings/schema_version", "roadmap.table.schema_version.invalid")
    if version != 2:
        report.add("table", "error", "roadmap.table.schema_version.unsupported", "/Settings/schema_version", "canonical table import requires schema_version 2")
    threshold = integer(settings.get("shift_threshold_days", "0"), report, "/Settings/shift_threshold_days", "roadmap.table.shift_threshold.invalid")
    model = compact({
        "schema_version": version,
        "title": settings.get("title", ""),
        "time_scale": settings.get("time_scale", "month"),
        "lane_dimension": settings.get("lane_dimension", "workstream"),
        "shift_threshold_days": threshold,
        "assumptions": split_ids(settings.get("assumptions", "")),
    })

    def entities(sheet, fields, list_fields=()):
        items = []
        seen = set()
        for row in tables[sheet]:
            iid = row.get("id", "")
            if not iid:
                report.add("table", "error", "roadmap.table.id.missing", f"/{sheet}/{row['_row']}/id", f"non-empty {sheet} row requires id")
                continue
            if iid in seen:
                report.add("table", "error", "roadmap.table.id.duplicate", f"/{sheet}/{row['_row']}/id", f"duplicate id {iid!r}")
            seen.add(iid)
            item = {field: split_ids(row.get(field, "")) if field in list_fields else row.get(field, "") for field in fields}
            items.append(compact(item))
        return sorted(items, key=lambda item: item.get("id", ""))

    model["lanes"] = entities("Lanes", ("id", "title", "description"))
    time_scale = model.get("time_scale", "month")
    task_fields = ["id", "title", "lane", "status", "owner", "risk", "milestones", "outcomes", "notes"]
    task_fields += ["start_order", "end_order"] if time_scale == "order" else ["start", "end"]
    model["tasks"] = entities("Tasks", tuple(task_fields), ("milestones", "outcomes"))
    for index, task in enumerate(model["tasks"]):
        if time_scale == "order":
            for field in ("start_order", "end_order"):
                if field in task:
                    task[field] = integer(task[field], report, f"/Tasks/{index}/{field}", "roadmap.table.order.invalid")
        else:
            for field in ("start", "end"):
                if field in task:
                    task[field] = iso_date(task[field], report, f"/Tasks/{index}/{field}")
    model["dependencies"] = entities("Dependencies", ("id", "from", "to", "type", "impact", "rationale"))
    model["outcomes"] = entities("Outcomes", ("id", "title", "description", "metric"))

    grouped = defaultdict(list)
    for row in tables["MilestoneHistory"]:
        mid = row.get("milestone_id", "")
        if not mid:
            report.add("table", "error", "roadmap.table.milestone_id.missing", f"/MilestoneHistory/{row['_row']}/milestone_id", "non-empty MilestoneHistory row requires milestone_id")
            continue
        grouped[mid].append(row)
    milestones = []
    coordinate_column = "planned_order" if time_scale == "order" else "planned_date"
    coordinate_key = "order" if time_scale == "order" else "date"
    for mid in sorted(grouped):
        rows = grouped[mid]
        revision_ids, revision_orders = set(), set()
        normalized_rows = []
        for row in rows:
            rid = row.get("revision_id", "")
            revision_order = integer(row.get("revision_order", ""), report, f"/MilestoneHistory/{row['_row']}/revision_order", "roadmap.table.revision_order.invalid")
            if rid in revision_ids:
                report.add("table", "error", "roadmap.table.revision_id.duplicate", f"/MilestoneHistory/{row['_row']}/revision_id", f"duplicate revision_id {rid!r} for milestone {mid!r}")
            if revision_order in revision_orders:
                report.add("table", "error", "roadmap.table.revision_order.duplicate", f"/MilestoneHistory/{row['_row']}/revision_order", f"duplicate revision_order {revision_order!r} for milestone {mid!r}")
            revision_ids.add(rid)
            revision_orders.add(revision_order)
            coordinate_value = row.get(coordinate_column, "")
            coordinate_value = integer(coordinate_value, report, f"/MilestoneHistory/{row['_row']}/{coordinate_column}", "roadmap.table.order.invalid") if time_scale == "order" else iso_date(coordinate_value, report, f"/MilestoneHistory/{row['_row']}/{coordinate_column}")
            recorded = row.get("recorded_at", "")
            revision = compact({
                "revision_id": rid, "revision_order": revision_order, "plan_version": row.get("plan_version", ""),
                coordinate_key: coordinate_value,
                "recorded_at": iso_date(recorded, report, f"/MilestoneHistory/{row['_row']}/recorded_at") if recorded else "",
                "reason": row.get("reason", ""), "_current": boolean(row.get("is_current", "")), "_row": row,
            })
            normalized_rows.append(revision)
        current_rows = [row for row in normalized_rows if row.get("_current")]
        if not current_rows:
            report.add("table", "error", "roadmap.table.current.missing", "/MilestoneHistory", f"milestone {mid!r} has no current revision")
            current = max(normalized_rows, key=lambda item: (item.get("revision_order", 0), item.get("revision_id", "")))
        elif len(current_rows) > 1:
            report.add("table", "error", "roadmap.table.current.multiple", "/MilestoneHistory", f"milestone {mid!r} has multiple current revisions")
            current = max(current_rows, key=lambda item: (item.get("revision_order", 0), item.get("revision_id", "")))
        else:
            current = current_rows[0]
        if any(item.get("revision_order", 0) > current.get("revision_order", 0) for item in normalized_rows if item is not current):
            report.add("table", "error", "roadmap.table.current.not_latest", "/MilestoneHistory", f"current revision for milestone {mid!r} must have the greatest revision_order")
        source_row = current["_row"]
        history = []
        for revision in sorted((item for item in normalized_rows if item is not current), key=lambda item: (item.get("revision_order", 0), item.get("revision_id", ""))):
            history.append({key: value for key, value in revision.items() if not key.startswith("_")})
        milestone = compact({
            "id": mid, "title": source_row.get("milestone_title", ""), "lane": source_row.get("lane", ""),
            coordinate_key: current.get(coordinate_key), "revision_id": current.get("revision_id"),
            "revision_order": current.get("revision_order"), "plan_version": current.get("plan_version"),
            "recorded_at": current.get("recorded_at", ""), "reason": current.get("reason", ""), "history": history,
            "status": source_row.get("status", ""), "owner": source_row.get("owner", ""), "risk": source_row.get("risk", ""),
            "outcomes": split_ids(source_row.get("outcomes", "")), "notes": source_row.get("notes", ""),
        })
        milestones.append(milestone)
    model["milestones"] = milestones

    normalized, source_report = validate_document(model, strict=strict)
    report.findings.extend(source_report["findings"])
    report.details["source"] = str(path.resolve())
    report.details["model"] = normalized if normalized is not None else model
    report.details["history_deltas"] = source_report.get("history_deltas", [])
    return model, report.finish(strict=False)


def main(argv=None):
    parser = argparse.ArgumentParser(description="Import canonical roadmap XLSX/CSV into strict roadmap.v2 YAML.")
    parser.add_argument("input", help="filled .xlsx or long-form .csv template")
    parser.add_argument("-o", "--output", required=True, help="output roadmap.v2 YAML path")
    parser.add_argument("--strict", action="store_true", help="treat schema and semantic warnings as errors")
    parser.add_argument("--json", action="store_true", help="print the machine-readable import report")
    parser.add_argument("--report", help="write the import report as JSON")
    args = parser.parse_args(argv)
    try:
        model, report = import_model(Path(args.input).expanduser().resolve(), strict=args.strict)
    except (OSError, ValueError, RuntimeError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1
    if args.report:
        report_path = Path(args.report).expanduser().resolve()
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    if args.json:
        print(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    if report["summary"]["errors"]:
        if not args.json:
            for finding in report["findings"]:
                if finding["severity"] == "error":
                    print(f"error: [{finding['code']}] {finding['path']}: {finding['message']}", file=sys.stderr)
        return 1
    output = Path(args.output).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(yaml.safe_dump(model, allow_unicode=True, sort_keys=False), encoding="utf-8")
    if not args.json:
        print(f"wrote roadmap.v2 YAML: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

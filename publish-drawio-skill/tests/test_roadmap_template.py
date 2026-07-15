import csv
import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
import unittest
import xml.etree.ElementTree as ET
import zipfile
from datetime import date
from pathlib import Path

import jsonschema
import openpyxl
import yaml


ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
ASSETS = ROOT / "assets" / "roadmap"
FIXTURES = ROOT / "tests" / "fixtures" / "roadmap"
XLSX = ASSETS / "roadmap-template.xlsx"
CSV = ASSETS / "roadmap-template.csv"
EXPECTED_SHEETS = [
    "Settings", "Lanes", "Tasks", "MilestoneHistory",
    "Dependencies", "Outcomes", "Lists", "Instructions",
]
EXPECTED_HEADERS = {
    "Settings": ["key", "value", "description"],
    "Lanes": ["id", "title", "description"],
    "Tasks": [
        "id", "title", "lane", "start", "end", "start_order", "end_order",
        "status", "owner", "risk", "milestones", "outcomes", "notes",
    ],
    "MilestoneHistory": [
        "milestone_id", "milestone_title", "lane", "revision_id",
        "revision_order", "plan_version", "planned_date", "planned_order",
        "recorded_at", "is_current", "reason", "status", "owner", "risk",
        "outcomes", "notes", "previous_planned_date", "shift_days",
        "cumulative_shift_days", "shift_state",
    ],
    "Dependencies": ["id", "from", "to", "type", "impact", "rationale"],
    "Outcomes": ["id", "title", "description", "metric"],
}


def digest(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def run(script, *args):
    return subprocess.run(
        [sys.executable, str(SCRIPTS / script), *map(str, args)],
        cwd=ROOT,
        text=True,
        capture_output=True,
    )


def cell_map(path):
    tree = ET.parse(path)
    return {cell.get("id"): cell for cell in tree.findall(".//mxCell") if cell.get("id")}


class CanonicalWorkbookContractTests(unittest.TestCase):
    def test_xlsx_sheets_headers_formulas_validations_and_sample_history(self):
        workbook = openpyxl.load_workbook(XLSX, data_only=False, keep_links=False)
        self.assertEqual(workbook.sheetnames, EXPECTED_SHEETS)
        for sheet, headers in EXPECTED_HEADERS.items():
            with self.subTest(sheet=sheet):
                self.assertEqual([cell.value for cell in workbook[sheet][1]], headers)

        history = workbook["MilestoneHistory"]
        self.assertEqual([history.cell(row, 5).value for row in range(2, 6)], [1, 2, 3, 4])
        dates = [history.cell(row, 7).value.date() for row in range(2, 6)]
        shifts = [(right - left).days for left, right in zip(dates, dates[1:])]
        self.assertEqual(shifts, [15, 14, -7])
        self.assertEqual((dates[-1] - dates[0]).days, 22)
        self.assertEqual([history.cell(row, 10).value for row in range(2, 6)], [False, False, False, True])
        for row in range(2, 6):
            for column in range(17, 21):
                with self.subTest(row=row, column=column):
                    self.assertEqual(history.cell(row, column).data_type, "f")
        self.assertIn("ShiftThresholdDays", history["T2"].value)
        self.assertGreaterEqual(len(workbook["Settings"].data_validations.dataValidation), 1)
        self.assertGreaterEqual(len(workbook["Tasks"].data_validations.dataValidation), 2)
        self.assertGreaterEqual(len(history.data_validations.dataValidation), 3)
        self.assertIsNone(workbook.vba_archive)
        self.assertFalse(getattr(workbook, "_external_links", []))
        workbook.close()

        with zipfile.ZipFile(XLSX) as archive:
            names = archive.namelist()
        self.assertFalse(any(name.endswith("vbaProject.bin") for name in names))
        self.assertFalse(any(name.startswith("xl/externalLinks/") for name in names))

    def test_xlsx_formula_table_and_validation_contract_extends_through_row_1001(self):
        workbook = openpyxl.load_workbook(XLSX, data_only=False, keep_links=False)
        history = workbook["MilestoneHistory"]
        table = history.tables["MilestoneHistoryTable"]
        self.assertEqual(table.ref, "A1:T1001")
        for row in range(2, 1002):
            with self.subTest(row=row):
                self.assertTrue(all(history.cell(row, column).data_type == "f" for column in range(17, 21)))
        for row in (2, 3, 1001):
            previous = history.cell(row, 17).value
            shift = history.cell(row, 18).value
            cumulative = history.cell(row, 19).value
            self.assertIn("MAXIFS", previous)
            self.assertIn('"<"&E', previous)
            self.assertNotIn(f"E{row}-1", previous)
            self.assertEqual(shift, f'=IF(OR(Q{row}="",G{row}=""),"",G{row}-Q{row})')
            self.assertIn("SUMIFS", cumulative)
            self.assertIn("$1001", previous)
            self.assertIn("$1001", cumulative)
        validation_ranges = {str(validation.sqref) for validation in history.data_validations.dataValidation}
        self.assertEqual(validation_ranges, {"C2:C1001", "J2:J1001", "L2:L1001"})
        workbook.close()

    def test_csv_is_long_form_data_only_fallback_with_matching_sample(self):
        with CSV.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        self.assertIn("entity_type", rows[0])
        self.assertTrue(set(["setting", "lane", "task", "milestone_revision", "dependency", "outcome"]).issubset({row["entity_type"] for row in rows}))
        revisions = [row for row in rows if row["entity_type"] == "milestone_revision"]
        self.assertEqual([row["revision_id"] for row in revisions], ["rev-1", "rev-2", "rev-3", "rev-4"])
        self.assertEqual([row["shift_days"] for row in revisions], ["", "15", "14", "-7"])
        self.assertEqual(revisions[-1]["cumulative_shift_days"], "22")
        self.assertFalse(any(value.lstrip().startswith("=") for row in rows for value in row.values() if value))


class TemplateCopyTests(unittest.TestCase):
    def test_copy_reports_absolute_path_hash_and_json_without_mutating_asset(self):
        original = digest(XLSX)
        with tempfile.TemporaryDirectory() as temp:
            proc = run("roadmap_template.py", temp, "--format", "xlsx", "--json")
            self.assertEqual(proc.returncode, 0, proc.stderr + proc.stdout)
            result = json.loads(proc.stdout)
            copied = Path(result["path"])
            self.assertTrue(copied.is_absolute())
            self.assertEqual(copied, Path(temp).resolve() / XLSX.name)
            self.assertEqual(result["sha256"], original)
            self.assertEqual(digest(copied), original)

            second = run("roadmap_template.py", temp, "--format", "xlsx", "--json")
            self.assertNotEqual(second.returncode, 0)
            self.assertEqual(json.loads(second.stdout)["code"], "roadmap.template.copy_failed")
            forced = run("roadmap_template.py", temp, "--format", "xlsx", "--force", "--json")
            self.assertEqual(forced.returncode, 0, forced.stderr + forced.stdout)
        self.assertEqual(digest(XLSX), original)


class RoadmapTableImportTests(unittest.TestCase):
    def import_workbook(self, mutate=None):
        temp = tempfile.TemporaryDirectory()
        source = Path(temp.name) / "source.xlsx"
        shutil.copyfile(XLSX, source)
        if mutate:
            workbook = openpyxl.load_workbook(source, keep_links=False)
            mutate(workbook)
            workbook.save(source)
            workbook.close()
        output = Path(temp.name) / "roadmap.yaml"
        report = Path(temp.name) / "report.json"
        proc = run("roadmap_table.py", source, "-o", output, "--strict", "--json", "--report", report)
        payload = json.loads(proc.stdout) if proc.stdout.strip() else json.loads(report.read_text(encoding="utf-8"))
        return temp, output, proc, payload

    def test_xlsx_imports_to_strict_v2_and_recalculates_untrusted_computed_cells(self):
        def mutate(workbook):
            history = workbook["MilestoneHistory"]
            for row in range(2, 6):
                history.cell(row, 17).value = "1900-01-01"
                history.cell(row, 18).value = 999999
                history.cell(row, 19).value = -999999
                history.cell(row, 20).value = "unchanged"

        temp, output, proc, report = self.import_workbook(mutate)
        try:
            self.assertEqual(proc.returncode, 0, proc.stderr + proc.stdout)
            model = yaml.safe_load(output.read_text(encoding="utf-8"))
            self.assertEqual(model["schema_version"], 2)
            self.assertEqual([item["revision_id"] for item in model["milestones"][0]["history"]], ["rev-1", "rev-2", "rev-3"])
            self.assertEqual(model["milestones"][0]["revision_id"], "rev-4")
            self.assertEqual([item["delta"] for item in report["history_deltas"]], [15, 14, -7])
            self.assertEqual([item["cumulative_delta"] for item in report["history_deltas"]], [15, 29, 22])
            validate = run("roadmap_validate.py", output, "--strict", "--json")
            self.assertEqual(validate.returncode, 0, validate.stderr + validate.stdout)
        finally:
            temp.cleanup()

    def test_duplicate_revision_multiple_current_invalid_date_and_order_fail_stably(self):
        cases = {
            "duplicate revision": (
                lambda wb: setattr(wb["MilestoneHistory"]["D3"], "value", "rev-1"),
                "roadmap.table.revision_id.duplicate",
            ),
            "multiple current": (
                lambda wb: setattr(wb["MilestoneHistory"]["J4"], "value", True),
                "roadmap.table.current.multiple",
            ),
            "invalid date": (
                lambda wb: setattr(wb["MilestoneHistory"]["G3"], "value", "2026-02-31"),
                "roadmap.table.date.invalid",
            ),
            "duplicate order": (
                lambda wb: setattr(wb["MilestoneHistory"]["E3"], "value", 1),
                "roadmap.table.revision_order.duplicate",
            ),
            "current is not latest": (
                lambda wb: (setattr(wb["MilestoneHistory"]["J4"], "value", True), setattr(wb["MilestoneHistory"]["J5"], "value", False)),
                "roadmap.table.current.not_latest",
            ),
        }
        for name, (mutate, code) in cases.items():
            with self.subTest(name=name):
                temp, output, proc, report = self.import_workbook(mutate)
                try:
                    self.assertNotEqual(proc.returncode, 0, proc.stderr + proc.stdout)
                    self.assertFalse(output.exists())
                    self.assertIn(code, {finding["code"] for finding in report["findings"]})
                finally:
                    temp.cleanup()

    def test_xlsx_nonempty_rows_require_ids_while_formula_only_rows_are_ignored(self):
        cases = {
            "entity id": (
                lambda wb: setattr(wb["Lanes"]["B3"], "value", "Missing ID"),
                "roadmap.table.id.missing",
            ),
            "milestone id": (
                lambda wb: setattr(wb["MilestoneHistory"]["B6"], "value", "Missing milestone ID"),
                "roadmap.table.milestone_id.missing",
            ),
        }
        for name, (mutate, expected_code) in cases.items():
            with self.subTest(name=name):
                temp, output, proc, report = self.import_workbook(mutate)
                try:
                    self.assertNotEqual(proc.returncode, 0, proc.stderr + proc.stdout)
                    self.assertFalse(output.exists())
                    self.assertIn(expected_code, {item["code"] for item in report["findings"]})
                finally:
                    temp.cleanup()

        # The bundled sheet intentionally contains formulas down to row 1001;
        # those formula-only rows are scaffolding, not empty entities.
        temp, output, proc, report = self.import_workbook()
        try:
            self.assertEqual(proc.returncode, 0, proc.stderr + proc.stdout)
            codes = {item["code"] for item in report["findings"]}
            self.assertNotIn("roadmap.table.id.missing", codes)
            self.assertNotIn("roadmap.table.milestone_id.missing", codes)
            self.assertEqual(len(yaml.safe_load(output.read_text(encoding="utf-8"))["milestones"]), 1)
        finally:
            temp.cleanup()

    def test_csv_nonempty_rows_require_ids_while_computed_only_rows_are_ignored(self):
        with CSV.open(encoding="utf-8-sig", newline="") as handle:
            fieldnames = csv.DictReader(handle).fieldnames
        cases = {
            "entity id": ({"entity_type": "lane", "title": "Missing ID"}, "roadmap.table.id.missing"),
            "milestone id": ({"entity_type": "milestone_revision", "milestone_title": "Missing milestone ID"}, "roadmap.table.milestone_id.missing"),
        }
        for name, (extra, expected_code) in cases.items():
            with self.subTest(name=name), tempfile.TemporaryDirectory() as temp:
                source = Path(temp) / "source.csv"
                shutil.copyfile(CSV, source)
                with source.open("a", encoding="utf-8", newline="") as handle:
                    writer = csv.DictWriter(handle, fieldnames=fieldnames)
                    writer.writerow(extra)
                output = Path(temp) / "roadmap.yaml"
                proc = run("roadmap_table.py", source, "-o", output, "--strict", "--json")
                self.assertNotEqual(proc.returncode, 0, proc.stderr + proc.stdout)
                report = json.loads(proc.stdout)
                self.assertIn(expected_code, {item["code"] for item in report["findings"]})

        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "source.csv"
            shutil.copyfile(CSV, source)
            with source.open("a", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writerow({"shift_days": "=999", "cumulative_shift_days": "=999"})
            output = Path(temp) / "roadmap.yaml"
            proc = run("roadmap_table.py", source, "-o", output, "--strict", "--json")
            self.assertEqual(proc.returncode, 0, proc.stderr + proc.stdout)
            self.assertEqual(len(yaml.safe_load(output.read_text(encoding="utf-8"))["milestones"]), 1)

    def test_sparse_revision_orders_import_in_sequence_with_correct_deltas(self):
        def mutate(workbook):
            history = workbook["MilestoneHistory"]
            history["E2"], history["E3"], history["E4"] = 1, 3, 5
            history["J4"] = True
            history["J5"] = False
            for column in range(1, 17):
                history.cell(5, column).value = None

        temp, output, proc, report = self.import_workbook(mutate)
        try:
            self.assertEqual(proc.returncode, 0, proc.stderr + proc.stdout)
            model = yaml.safe_load(output.read_text(encoding="utf-8"))
            milestone = model["milestones"][0]
            self.assertEqual([item["revision_order"] for item in milestone["history"]], [1, 3])
            self.assertEqual(milestone["revision_order"], 5)
            self.assertEqual([item["delta"] for item in report["history_deltas"]], [15, 14])
            self.assertEqual([item["cumulative_delta"] for item in report["history_deltas"]], [15, 29])
        finally:
            temp.cleanup()

    def test_csv_import_matches_canonical_history(self):
        with tempfile.TemporaryDirectory() as temp:
            output = Path(temp) / "roadmap.yaml"
            proc = run("roadmap_table.py", CSV, "-o", output, "--strict", "--json")
            self.assertEqual(proc.returncode, 0, proc.stderr + proc.stdout)
            model = yaml.safe_load(output.read_text(encoding="utf-8"))
            self.assertEqual(model["milestones"][0]["date"], "2026-10-07")
            self.assertEqual(len(model["milestones"][0]["history"]), 3)


class RoadmapV2HistoryTests(unittest.TestCase):
    def test_v2_schema_compiles_and_full_history_strictly_validates(self):
        schema = json.loads((ROOT / "data" / "roadmap.v2.schema.json").read_text(encoding="utf-8"))
        jsonschema.Draft202012Validator.check_schema(schema)
        proc = run("roadmap_validate.py", FIXTURES / "full_history.yaml", "--strict", "--json")
        self.assertEqual(proc.returncode, 0, proc.stderr + proc.stdout)
        deltas = json.loads(proc.stdout)["history_deltas"]
        self.assertEqual([(item["delta"], item["cumulative_delta"]) for item in deltas], [(15, 15), (14, 29), (-7, 22)])

    def test_full_history_generates_three_edges_profiles_and_is_deterministic(self):
        source = FIXTURES / "full_history.yaml"
        with tempfile.TemporaryDirectory() as temp:
            artifact = Path(temp) / "roadmap.drawio"
            generation = run("roadmap.py", source, "-o", artifact)
            self.assertEqual(generation.returncode, 0, generation.stderr + generation.stdout)
            cells = cell_map(artifact)
            expected_edges = [
                "history_shift_m-wallet-pilot_rev-1_rev-2",
                "history_shift_m-wallet-pilot_rev-2_rev-3",
                "history_shift_m-wallet-pilot_rev-3_rev-4",
            ]
            self.assertEqual([cells[cid].get("value") for cid in expected_edges], ["+15d", "+14d", "-7d"])
            self.assertTrue({"history_m-wallet-pilot_rev-1", "history_m-wallet-pilot_rev-2", "history_m-wallet-pilot_rev-3"}.issubset(cells))
            profile = run("validate.py", artifact, "--profile", "roadmap", "--source", source, "--strict", "--json")
            self.assertEqual(profile.returncode, 0, profile.stderr + profile.stdout)
            deterministic = run("verify_determinism.py", "roadmap", source, "--json")
            self.assertEqual(deterministic.returncode, 0, deterministic.stderr + deterministic.stdout)

    def test_source_aware_validation_catches_missing_history_marker_and_edge(self):
        source = FIXTURES / "full_history.yaml"
        cases = {
            "history_m-wallet-pilot_rev-2": "artifact.coverage.milestone_history",
            "history_shift_m-wallet-pilot_rev-2_rev-3": "artifact.coverage.history_shift",
        }
        for remove_id, expected_code in cases.items():
            with self.subTest(remove_id=remove_id), tempfile.TemporaryDirectory() as temp:
                artifact = Path(temp) / "roadmap.drawio"
                self.assertEqual(run("roadmap.py", source, "-o", artifact).returncode, 0)
                tree = ET.parse(artifact)
                root = tree.find(".//root")
                root.remove(next(cell for cell in root.findall("mxCell") if cell.get("id") == remove_id))
                tree.write(artifact, encoding="utf-8", xml_declaration=True)
                profile = run("validate.py", artifact, "--profile", "roadmap", "--source", source, "--json")
                self.assertNotEqual(profile.returncode, 0)
                codes = {item["code"] for item in json.loads(profile.stdout)["findings"]}
                self.assertIn(expected_code, codes)

    def test_source_aware_validation_catches_current_milestone_metadata_tamper(self):
        source = FIXTURES / "full_history.yaml"
        with tempfile.TemporaryDirectory() as temp:
            artifact = Path(temp) / "roadmap.drawio"
            generation = run("roadmap.py", source, "-o", artifact)
            self.assertEqual(generation.returncode, 0, generation.stderr + generation.stdout)
            tree = ET.parse(artifact)
            current = next(cell for cell in tree.findall(".//mxCell") if cell.get("id") == "milestone_m-wallet-pilot")
            current.set("data-plan-version", "tampered")
            tree.write(artifact, encoding="utf-8", xml_declaration=True)
            profile = run("validate.py", artifact, "--profile", "roadmap", "--source", source, "--json")
            self.assertNotEqual(profile.returncode, 0)
            codes = {item["code"] for item in json.loads(profile.stdout)["findings"]}
            self.assertIn("artifact.coverage.current_plan_version", codes)


class RoadmapWorkflowDocumentationTests(unittest.TestCase):
    def test_docs_require_canonical_working_copy_confirmation_gates(self):
        skill = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        reference = (ROOT / "references" / "roadmap.md").read_text(encoding="utf-8")
        combined = skill + "\n" + reference
        self.assertIn("canonical XLSX", combined)
        self.assertIn("CSV fallback", combined)
        self.assertIn("working directory", combined)
        self.assertIn("Never", combined)
        self.assertIn("bundled", combined)
        self.assertIn("STOP until the user confirms", skill)
        self.assertIn("fill only the working copy", reference)
        self.assertIn("until the user confirms generation", reference)
        self.assertIn("summarize", reference.lower())
        self.assertIn("must not be blocked", skill)
        self.assertIn("python3 -m pip install -r <this-skill-dir>/requirements.txt", combined)


if __name__ == "__main__":
    unittest.main()
